#!/usr/bin/env python3
"""
Step 5 (deterministic, no LLM) for forward-guidance-extractor.

Builds the final .xlsx from the SAME JSON DTOs that were persisted to the DB
via save_forward_guidance.js -- the workbook is a template render of that data,
never an independent source (skills/_shared/conventions.md #5, JSON-first).

Input: a JSON array of report DTOs (the `{id, companyId, quarter, guidance:[...],
transcriptAvailable, staleGuidanceNote, ...}` objects `save_forward_guidance.js`
printed / that `db.get('reports', id)` would return), plus a `missing` array
(companies with no usable transcript at all, from classify_transcript_status.py).

Usage:
    python3 build_guidance_workbook.py \
        --dtos company_dtos.json \
        --missing missing_companies.json \
        --out /path/to/Forward_Guidance_<date>.xlsx

Requires: openpyxl (pip install openpyxl --break-system-packages)
"""
import argparse
import json

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
FLAG_FILL = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")

GUIDANCE_COLUMNS = [
    ("Company", 22),
    ("Ticker", 14),
    ("Quarter (Concall)", 14),
    ("Metric Category", 16),
    ("Metric", 20),
    ("Period Guided", 14),
    ("Guidance (Absolute (Relative %))", 30),
    ("Base Period Referenced", 16),
    ("Management Quote", 60),
    ("Source", 12),
    ("Derived Field", 12),
    ("Stale Guidance Flag", 16),
]

MISSING_COLUMNS = [("Company/Ticker", 22), ("Quarter", 14), ("Reason", 50)]

# Optional scan-table columns, present when the run's input was a Stockscans
# saved-scan URL -- read from each DTO's `scanRow` (carried through from
# guidance-document-extractor, never re-fetched here). Only the columns
# actually present on at least one DTO are added, in this preferred order.
SCAN_COL_ORDER = [
    "Market Capitalization", "Market Cap",
    "Price To Earnings", "P/E", "PE",
    "CFO To PAT", "CFO/PAT",
    "Change In FII Holdings Latest Quarter", "Change in FII Holdings Latest Quarter",
    "FII Holdings",
]


def _scan_columns_present(dtos):
    seen = []
    for d in dtos:
        for k in (d.get("scanRow") or {}).keys():
            if k in SCAN_COL_ORDER and k not in seen:
                seen.append(k)
    return [c for c in SCAN_COL_ORDER if c in seen]


def _write_header(ws, columns):
    for idx, (title, width) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=idx, value=title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A2"


def build_guidance_sheet(ws, dtos, scan_cols):
    _write_header(ws, GUIDANCE_COLUMNS + [(c, 16) for c in scan_cols])
    row = 2
    for dto in dtos:
        company_name = dto.get("companyName") or dto.get("companyId")
        stale = dto.get("staleGuidanceNote")
        scan_row = dto.get("scanRow") or {}
        scan_vals = [scan_row.get(c, "") for c in scan_cols]
        for item in dto.get("guidance", []):
            values = [
                company_name,
                dto.get("companyId"),
                dto.get("quarter"),
                item.get("metric_category"),
                item.get("metric"),
                item.get("period_guided"),
                item.get("display") or "",
                item.get("base_period") or "",
                item.get("quote") or "",
                item.get("source") or "",
                item.get("derived_field") or "none",
                stale or "",
            ] + scan_vals
            for col_idx, val in enumerate(values, start=1):
                cell = ws.cell(row=row, column=col_idx, value=val)
                if stale and col_idx == len(GUIDANCE_COLUMNS):
                    cell.fill = FLAG_FILL
            row += 1
    return row - 2  # data rows written


def build_missing_sheet(ws, missing, scan_cols):
    _write_header(ws, MISSING_COLUMNS + [(c, 16) for c in scan_cols])
    row = 2
    for m in missing:
        ws.cell(row=row, column=1, value=m.get("ticker"))
        ws.cell(row=row, column=2, value=m.get("quarter"))
        ws.cell(row=row, column=3, value=m.get("reason"))
        scan_row = m.get("scanRow") or {}
        for j, c in enumerate(scan_cols, 4):
            ws.cell(row=row, column=j, value=scan_row.get(c, ""))
        for c in range(1, 4):
            ws.cell(row=row, column=c).fill = FLAG_FILL
        row += 1
    return row - 2


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dtos", required=True, help="JSON array of forward-guidance report DTOs")
    ap.add_argument("--missing", help="JSON array of {ticker, quarter, reason} with no usable transcript")
    ap.add_argument("--out", required=True, help="output .xlsx path")
    args = ap.parse_args()

    dtos = json.load(open(args.dtos))
    missing = json.load(open(args.missing)) if args.missing else []

    scan_cols = _scan_columns_present(dtos + missing)

    wb = Workbook()
    ws_guidance = wb.active
    ws_guidance.title = "Forward Guidance"
    n_guidance_rows = build_guidance_sheet(ws_guidance, dtos, scan_cols)

    ws_missing = wb.create_sheet("Missing Transcripts")
    n_missing_rows = build_missing_sheet(ws_missing, missing, scan_cols)

    wb.save(args.out)
    print(json.dumps({
        "path": args.out,
        "guidance_rows": n_guidance_rows,
        "missing_rows": n_missing_rows,
        "companies_covered": len({d.get("companyId") for d in dtos}),
    }, indent=2))


if __name__ == "__main__":
    main()
