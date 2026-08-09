#!/usr/bin/env python3
"""
Step 4 (deterministic, no LLM) for pead-surprise-ranker.

Renders the ranked/scored companies (compute_pead_score.py output) plus the
excluded/no-visibility list plus a methodology note into one .xlsx workbook.
Pure template render -- the JSON produced by Steps 1-3 is the source of
truth; this script must never be the place new judgment gets added.

Usage:
  python3 build_pead_workbook.py \
    --ranked ranked.json \
    --excluded excluded.json \
    --methodology methodology.txt \
    --out PEAD_Ranking_<date>.xlsx \
    [--guidance-dtos forward_guidance_dtos.json]

`--guidance-dtos` (added 2026-08-09, so the user only has to look at one
sheet instead of cross-referencing the separate Forward Guidance workbook):
an array of the underlying `forward-guidance` report DTOs this ranking was
built from (the same `{companyId, quarter, guidance:[...]}` objects
`forward-guidance-extractor`'s Phase 3/4 already assembles). When supplied,
the "PEAD Ranking" sheet becomes ONE ROW PER GUIDANCE LINE ITEM (company-level
columns -- Rank, Score, Thesis, etc. -- repeat down every row for that
company) rather than one row per company, and gains every column the
standalone Forward Guidance workbook has (Metric Category, Metric, Period
Guided, Guidance display, Base Period, Management Quote, Source, Derived
Field, Stale flag). Companies with no matching DTO (or none passed) fall back
to the prior one-row-per-company rendering with guidance columns left blank --
never an error, since not every PEAD run necessarily has a paired
forward-guidance batch to hand.
"""
import argparse
import json

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TIER_FILL = {1: "C6E0B4", 2: "FFF2CC", 3: "FCE4D6", 4: "F8CBAD"}


# When the run's input was a Stockscans saved-scan URL, each ranked/excluded
# record may carry a `scan_cols` dict (Market Cap, P/E, CFO/PAT, "Change in
# FII Holdings Latest Quarter", FII Holdings, etc. -- lifted straight from
# the scanRow already fetched by guidance-document-extractor / carried
# through forward-guidance-extractor, never re-fetched here). Columns are
# added dynamically, driven by whatever keys are actually present, so this
# script stays agnostic to the exact scan-column set rather than hardcoding
# Stockscans' current column names.
SCAN_COL_ORDER = [
    "Market Capitalization", "Market Cap",
    "Price To Earnings", "P/E", "PE",
    "CFO To PAT", "CFO/PAT",
    "Change In FII Holdings Latest Quarter", "Change in FII Holdings Latest Quarter",
    "FII Holdings",
]


def _scan_columns_present(records):
    seen = []
    for r in records:
        for k in (r.get("scan_cols") or {}).keys():
            if k not in seen:
                seen.append(k)
    # Prefer the canonical order above when present, then append anything
    # else observed (keeps output deterministic without silently dropping
    # unexpected scan columns).
    ordered = [c for c in SCAN_COL_ORDER if c in seen]
    ordered += [c for c in seen if c not in ordered]
    return ordered


# Forward-guidance line-item columns, matching forward-guidance-extractor's
# own build_guidance_workbook.py GUIDANCE_COLUMNS 1:1 (minus Company/Ticker/
# Quarter, which are already the PEAD sheet's own leading columns) -- kept in
# sync deliberately so a user who has seen the standalone Forward Guidance
# workbook recognises the same fields here.
GUIDANCE_ITEM_HEADERS = [
    "Metric Category", "Metric", "Period Guided",
    "Guidance (Absolute (Relative %))", "Base Period Referenced",
    "Management Quote", "Guidance Source", "Derived Field",
]


def _guidance_item_row(item):
    return [
        item.get("metric_category", ""),
        item.get("metric", ""),
        item.get("period_guided", ""),
        item.get("display", ""),
        item.get("base_period", "") or "",
        item.get("quote", ""),
        item.get("source", ""),
        item.get("derived_field", ""),
    ]


def build_ranked_sheet(ws, ranked, guidance_by_ticker=None):
    guidance_by_ticker = guidance_by_ticker or {}
    scan_cols = _scan_columns_present(ranked)
    company_headers = [
        "Rank", "Ticker", "Company", "Sector", "Visibility Tier", "Composite Score",
        "Revenue Guidance", "Margin / Margin Direction", "PAT Lever",
        "Evidence Strength", "Thesis / Why it might beat", "Key Assumptions (explicit)",
        "Score Breakdown",
    ]
    has_guidance = any(r["ticker"] in guidance_by_ticker for r in ranked)
    headers = company_headers + scan_cols + (GUIDANCE_ITEM_HEADERS if has_guidance else [])
    n_company_cols = len(company_headers) + len(scan_cols)
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(wrap_text=True, vertical="top")

    row = 2
    n_data_rows = 0
    for rank, r in enumerate(ranked, 1):
        company_vals = [
            rank, r["ticker"], r.get("name", ""), r.get("sector", ""), r.get("tier"),
            r.get("composite_score"),
            r.get("rev_guided") or "(none disclosed)",
            f"{r.get('margin_guided') or '(none disclosed)'}  [{r.get('margin_dir')}]",
            r.get("pat_lever"),
            r.get("evidence"),
            r.get("thesis"),
            "; ".join(r.get("assumptions") or []),
            " | ".join(r.get("score_breakdown") or []),
        ] + [(r.get("scan_cols") or {}).get(c, "") for c in scan_cols]

        items = guidance_by_ticker.get(r["ticker"]) or []
        # One row per guidance line item, company columns repeated down each
        # row -- if there's no matching forward-guidance DTO for this ticker,
        # still emit exactly one row (company columns only, guidance columns
        # blank) so every ranked company appears at least once.
        rows_for_company = items if items else [None]
        for item in rows_for_company:
            vals = company_vals + (_guidance_item_row(item) if item else [""] * len(GUIDANCE_ITEM_HEADERS))
            for i, v in enumerate(vals, 1):
                if not has_guidance and i > n_company_cols:
                    continue
                cell = ws.cell(row=row, column=i, value=v)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                if i == 5:
                    fill = TIER_FILL.get(r.get("tier"), "FFFFFF")
                    cell.fill = PatternFill(start_color=fill, end_color=fill, fill_type="solid")
            row += 1
            n_data_rows += 1

    widths = [5, 14, 26, 20, 8, 9, 32, 32, 20, 12, 45, 40, 45] + [16] * len(scan_cols)
    if has_guidance:
        widths += [16, 20, 14, 26, 16, 55, 12, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    return n_data_rows


def build_excluded_sheet(ws, excluded):
    scan_cols = _scan_columns_present(excluded)
    headers = ["Ticker", "Why excluded from ranking", "Base-fundamentals note (NOT a forecast)"] + scan_cols
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(wrap_text=True)
    row = 2
    for e in excluded:
        ws.cell(row=row, column=1, value=e.get("ticker"))
        c2 = ws.cell(row=row, column=2, value=e.get("reason", ""))
        c2.alignment = Alignment(wrap_text=True)
        c3 = ws.cell(row=row, column=3, value=e.get("note", ""))
        c3.alignment = Alignment(wrap_text=True)
        for j, c in enumerate(scan_cols, 4):
            ws.cell(row=row, column=j, value=(e.get("scan_cols") or {}).get(c, ""))
        row += 1
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 55
    for j, c in enumerate(scan_cols, 4):
        ws.column_dimensions[get_column_letter(j)].width = 16
    return row - 2


def build_methodology_sheet(ws, lines):
    for i, line in enumerate(lines, 1):
        c = ws.cell(row=i, column=1, value=line)
        if i == 1:
            c.font = Font(bold=True, size=13)
        c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 140


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--ranked", required=True)
    ap.add_argument("--excluded", required=True)
    ap.add_argument("--methodology", required=True, help="Plain-text file, one methodology line per row")
    ap.add_argument("--out", required=True)
    ap.add_argument(
        "--guidance-dtos",
        required=False,
        help="Optional: array of forward-guidance report DTOs ({companyId, guidance:[...]}) "
        "to fold each company's guidance line items directly into the PEAD Ranking sheet.",
    )
    args = ap.parse_args()

    ranked = json.load(open(args.ranked))
    excluded = json.load(open(args.excluded))
    methodology_lines = [l.rstrip("\n") for l in open(args.methodology)]

    guidance_by_ticker = {}
    if args.guidance_dtos:
        dtos = json.load(open(args.guidance_dtos))
        for dto in dtos:
            ticker = dto.get("companyId")
            if ticker:
                guidance_by_ticker[ticker] = dto.get("guidance") or []

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "PEAD Ranking"
    n_ranked = build_ranked_sheet(ws1, ranked, guidance_by_ticker)

    ws2 = wb.create_sheet("No Visibility (Excluded)")
    n_excluded = build_excluded_sheet(ws2, excluded)

    ws3 = wb.create_sheet("Methodology & Caveats")
    build_methodology_sheet(ws3, methodology_lines)

    wb.save(args.out)
    print(json.dumps({"path": args.out, "ranked_rows": n_ranked, "excluded_rows": n_excluded}))


if __name__ == "__main__":
    main()
